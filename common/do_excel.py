# -*- coding: utf-8 -*-
# @Time    : 2019/11/21 0021 下午 11:48
# @Author  : wulin
# @Email   : 286075568@qq.com
# @FileName: do_excel.py
# @Software: PyCharm


# excel的改造升级

from openpyxl import *
import os, traceback
from copy import copy

from openpyxl.utils import get_column_letter


class Reader:
    def __init__(self):
        # 整个excel工作簿缓存
        self.workbook = None
        # 当前工作sheet
        self.sheet = None
        self.rows = 0
        self.r = 0

    # 读取excel
    def open_ecxel(self, file_path):
        # 如果打开的文件不存在，就报错
        if not os.path.isfile(file_path):
            print("error：%s not exist!" % (file_path))
            return
        self.wb = load_workbook(file_path)

        sheets = self.wb.sheetnames

        self.sheet = sheets[0]
        # self.rows = self.sheet.max_row
        # 设置默认读取为第一行
        self.r = 0

        return sheets

    # 获取sheet页面
    def get_sheets(self):
        # 获取所有sheet的名字，并返回为一个列表
        # sheets = self.wb.get_sheet_names()
        sheets = self.wb.sheetnames
        # print(sheets)
        return sheets

    # 切换sheet页面
    def set_sheet(self, name):
        self.sheet = self.wb[name]
        self.rows = self.sheet.max_row

        self.r = 0
        return

    # 读取
    def readline(self, row):

        if self.r < self.rows:

            # 获取最大的列数
            columns = self.sheet.max_column
            rows = self.sheet.max_row
            rowdata = []
            # 循环遍历

            for i in range(1, columns + 1):
                cellvalue = self.sheet.cell(row=row + 1, column=i).value

                if cellvalue != None:
                    rowdata.append(cellvalue)
                else:
                    rowdata.append('')

            return rowdata


class Writer:

    def __init__(self):
        # 读取需要复制的excel
        self.workbook = None
        # 拷贝的工作空间
        self.wb = None
        # 当前工作的sheet页
        self.sheet = None
        # 记录生成的文件，用来保存
        self.df = None
        # 记录写入的行
        self.row = 0
        # 记录写入的列
        self.clo = 0

    # 复制并打开excel
    def copy_open(self, srcfile, dstfile):
        # 判断要复制的文件是否存在
        if not os.path.isfile(srcfile):
            print(srcfile + " not exist!")
            return

        # 判断要新建的文档是否存在，存在则提示
        if os.path.isfile(dstfile):
            print("warning：" + dstfile + " file already exist!")

        # 记录要保存的文件
        self.df = dstfile
        self.workbook = load_workbook(filename=srcfile)
        # 拷贝
        self.wb = copy(self.workbook)

        # 默认使用第一个sheet
        # sheet = wb.get_sheet('Sheet1')
        return

    # 获取sheet页面
    def get_sheets(self):
        # 获取所有sheet的名字，并返回为一个列表
        sheets = self.wb.sheetnames
        print(sheets)
        return sheets

    # 切换sheet页面
    def set_sheet(self, name):
        # 通过sheet名字，切换sheet页面
        # self.sheet = self.wb.get_sheet(name)
        self.sheet = self.wb[name]

        return

    # 写入指定单元格，保留原格式
    def write(self, r, c, value):
        # 获取要写入的单元格
        def _getCell(sheet, r, c):
            """ HACK: Extract the internal xlwt cell representation. """

        self.sheet.cell(r, c, value)

        return

    # 保存
    def save_close(self):
        # 保存复制后的文件到硬盘
        try:
            self.wb.save(self.df)
        except Exception as e:
            print("error：文件保存失败！")
            # print(traceback.format_exc(e))

        return
